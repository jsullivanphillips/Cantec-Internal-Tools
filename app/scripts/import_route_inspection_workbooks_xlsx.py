from __future__ import annotations

import argparse
import csv
import logging
from dataclasses import dataclass
from datetime import date, datetime, time
from io import StringIO
from pathlib import Path

import openpyxl
from sqlalchemy import func

from app import create_app
from app.db_models import MonthlyRoute, MonthlyRouteRun, db
from app.monthly.route_inspection_csv_import import (
    ImportResult,
    _find_data_header_row_index,
    parse_sheet_meta,
    run_route_inspection_csv_import,
)
from app.monthly.runs import get_or_create_monthly_route_run

LOG = logging.getLogger("import_route_inspection_workbooks_xlsx")


def _configure_logging() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")


def _month_first(d: date) -> date:
    return date(d.year, d.month, 1)


def _cell_to_csv_string(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, time):
        # Keep the ':' so analyze_sheet_time_cells recognizes a clock.
        return value.strftime("%H:%M")
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _worksheet_to_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[list[str]]:
    rows: list[list[str]] = []
    for tup in ws.iter_rows(values_only=True):
        # Keep trailing empties trimmed; CSV importer is robust to ragged rows.
        out = [_cell_to_csv_string(v) for v in tup]
        while out and out[-1] == "":
            out.pop()
        rows.append(out)
    return rows


def _rows_to_csv_bytes(rows: list[list[str]]) -> bytes:
    buf = StringIO(newline="")
    w = csv.writer(buf)
    for r in rows:
        w.writerow(r)
    return buf.getvalue().encode("utf-8")

def _write_csv(path: Path, headers: list[str], rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def _next_monthly_route_run_id() -> int:
    current = db.session.query(func.coalesce(func.max(MonthlyRouteRun.id), 0)).scalar()
    return int(current or 0) + 1


def _get_or_create_run_no_commit(route_id: int, month_first: date) -> MonthlyRouteRun:
    existing = MonthlyRouteRun.query.filter_by(
        monthly_route_id=route_id, month_date=month_first
    ).one_or_none()
    if existing is not None:
        return existing
    run = MonthlyRouteRun(
        id=_next_monthly_route_run_id(),
        monthly_route_id=route_id,
        month_date=month_first,
        opened_at=datetime.now(),
        started_at=None,
        status="open",
        source="xlsx_bulk_import",
    )
    db.session.add(run)
    return run


@dataclass
class SheetCandidate:
    workbook_path: Path
    sheet_name: str
    route_number: int | None
    month_date: date | None
    route_label: str | None


def _inspect_sheet_candidate(
    workbook_path: Path,
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> SheetCandidate | None:
    rows = _worksheet_to_rows(ws)
    try:
        hdr_idx = _find_data_header_row_index(rows)
    except ValueError:
        return None
    rn, md, lab = parse_sheet_meta(rows, hdr_idx)
    return SheetCandidate(
        workbook_path=workbook_path,
        sheet_name=ws.title,
        route_number=rn,
        month_date=md,
        route_label=lab,
    )


def _iter_workbooks(dir_path: Path) -> list[Path]:
    if not dir_path.exists():
        raise SystemExit(f"Directory not found: {dir_path}")
    if not dir_path.is_dir():
        raise SystemExit(f"Not a directory: {dir_path}")
    out: list[Path] = []
    for p in sorted(dir_path.glob("*.xlsx")):
        if p.name.startswith("~$"):
            continue
        out.append(p)
    return out


def run_bulk_import(
    *,
    workbooks_dir: Path,
    min_month: date,
    commit: bool,
    allow_completed_runs: bool,
    only_workbook: str | None,
    max_sheets: int | None,
) -> None:
    mode = "commit" if commit else "dry-run (no DB commit)"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    logs_dir = Path("logs")
    issues_csv_path = logs_dir / f"xlsx_bulk_route_run_issues_{timestamp}.csv"
    sheet_summary_csv_path = logs_dir / f"xlsx_bulk_route_run_sheet_summary_{timestamp}.csv"

    print(f"[xlsx-bulk] Starting {mode}.", flush=True)
    print(f"[xlsx-bulk] Workbooks dir: {workbooks_dir}", flush=True)
    print(f"[xlsx-bulk] Min month: {min_month.isoformat()} (inclusive)", flush=True)
    print(f"[xlsx-bulk] Issues log: {issues_csv_path}", flush=True)

    workbook_paths = _iter_workbooks(workbooks_dir)
    if only_workbook:
        workbook_paths = [p for p in workbook_paths if p.name == only_workbook]
        if not workbook_paths:
            raise SystemExit(f"No workbook named {only_workbook!r} found in {workbooks_dir}")

    total_imported = 0
    total_skipped_old = 0
    total_skipped_unreadable = 0
    total_skipped_no_meta = 0
    total_skipped_route_missing = 0
    total_skipped_completed = 0
    total_issues = 0

    issue_rows: list[dict[str, object]] = []
    sheet_rows: list[dict[str, object]] = []

    for wb_path in workbook_paths:
        print(f"[xlsx-bulk] Workbook: {wb_path.name}", flush=True)
        try:
            wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
        except Exception as e:
            total_skipped_unreadable += 1
            LOG.warning("Skipping unreadable workbook %s: %s", wb_path, e)
            continue

        for sheet_name in wb.sheetnames:
            if max_sheets is not None and total_imported >= max_sheets:
                break

            ws = wb[sheet_name]
            candidate = _inspect_sheet_candidate(wb_path, ws)
            if candidate is None:
                total_skipped_no_meta += 1
                continue

            if candidate.month_date is None or candidate.route_number is None:
                total_skipped_no_meta += 1
                continue

            month_first = _month_first(candidate.month_date)
            if month_first < min_month:
                total_skipped_old += 1
                continue

            route = (
                db.session.query(MonthlyRoute)
                .filter(MonthlyRoute.route_number == int(candidate.route_number))
                .one_or_none()
            )
            if route is None:
                total_skipped_route_missing += 1
                LOG.warning(
                    "Skipping sheet %s / %s: route_number=%s not found in MonthlyRoute.",
                    wb_path.name,
                    sheet_name,
                    candidate.route_number,
                )
                continue

            if commit:
                run = get_or_create_monthly_route_run(
                    int(route.id), month_first, source="xlsx_bulk_import"
                )
            else:
                run = _get_or_create_run_no_commit(int(route.id), month_first)

            if not allow_completed_runs and getattr(run, "completed_at", None) is not None:
                total_skipped_completed += 1
                continue

            rows = _worksheet_to_rows(ws)
            csv_bytes = _rows_to_csv_bytes(rows)

            try:
                result: ImportResult = run_route_inspection_csv_import(
                    csv_bytes=csv_bytes,
                    run=run,
                    route=route,
                    month_date=month_first,
                    dry_run=not commit,
                )
            except Exception as e:
                db.session.rollback()
                total_skipped_unreadable += 1
                LOG.warning(
                    "Import failed for %s / %s: %s",
                    wb_path.name,
                    sheet_name,
                    e,
                )
                continue

            total_imported += 1
            total_issues += len(result.issues)
            sheet_rows.append(
                {
                    "workbook": wb_path.name,
                    "sheet": sheet_name,
                    "month_date": month_first.isoformat(),
                    "route_number": int(candidate.route_number),
                    "route_id": int(route.id),
                    "run_id": int(run.id),
                    "issues": len(result.issues),
                    "locations_updated": int(result.locations_updated),
                    "history_upserts": int(result.history_upserts),
                }
            )

            for i in result.issues:
                issue_rows.append(
                    {
                        "workbook": wb_path.name,
                        "sheet": sheet_name,
                        "month_date": month_first.isoformat(),
                        "route_number": int(candidate.route_number),
                        "route_id": int(route.id),
                        "run_id": int(run.id),
                        "issue_kind": i.kind,
                        "sheet_row": int(i.csv_row),
                        "detail": i.detail,
                    }
                )

            if len(result.issues) > 0:
                print(
                    f"[xlsx-bulk] Imported {wb_path.name} / {sheet_name} "
                    f"({month_first.isoformat()}) with {len(result.issues)} issue(s).",
                    flush=True,
                )
            else:
                print(
                    f"[xlsx-bulk] Imported {wb_path.name} / {sheet_name} ({month_first.isoformat()}).",
                    flush=True,
                )

        # Close file handles in read_only mode.
        try:
            wb.close()
        except Exception:
            pass

    if not commit:
        # Defensive: ensure no accidental persistence from intermediate commits.
        db.session.rollback()

    _write_csv(
        issues_csv_path,
        [
            "workbook",
            "sheet",
            "month_date",
            "route_number",
            "route_id",
            "run_id",
            "issue_kind",
            "sheet_row",
            "detail",
        ],
        issue_rows,
    )
    _write_csv(
        sheet_summary_csv_path,
        [
            "workbook",
            "sheet",
            "month_date",
            "route_number",
            "route_id",
            "run_id",
            "issues",
            "locations_updated",
            "history_upserts",
        ],
        sheet_rows,
    )

    print(
        "[xlsx-bulk] Summary — "
        f"imported_sheets: {total_imported}, "
        f"skipped_old: {total_skipped_old}, "
        f"skipped_no_meta: {total_skipped_no_meta}, "
        f"skipped_route_missing: {total_skipped_route_missing}, "
        f"skipped_completed: {total_skipped_completed}, "
        f"skipped_unreadable: {total_skipped_unreadable}, "
        f"total_issues: {total_issues}. "
        f"Wrote: {issues_csv_path} and {sheet_summary_csv_path}",
        flush=True,
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Bulk import technician run worksheets from .xlsx workbooks into monthly route runs."
    )
    parser.add_argument(
        "--workbooks-dir",
        default=r"C:\Users\jamie\Cantec Fire Alarms\Cantec Office - Documents\Cantec\Operations\Job Coordination\Monthly Inspections\!Monthly Bell Testing Routes",
        help="Directory containing route workbooks (.xlsx).",
    )
    parser.add_argument(
        "--min-month",
        default="2025-12-01",
        help="Only import worksheets with DATE month >= this (YYYY-MM-DD, first-of-month recommended).",
    )
    parser.add_argument(
        "--commit",
        action="store_true",
        help="Persist changes. If omitted, runs in dry-run mode.",
    )
    parser.add_argument(
        "--allow-completed-runs",
        action="store_true",
        help="Allow importing into runs already marked completed (not recommended).",
    )
    parser.add_argument(
        "--only-workbook",
        default=None,
        help="Only process a single workbook filename (exact match).",
    )
    parser.add_argument(
        "--max-sheets",
        type=int,
        default=None,
        help="Stop after importing N worksheets (for testing).",
    )
    return parser.parse_args()


def main() -> None:
    _configure_logging()
    args = parse_args()

    try:
        min_month = date.fromisoformat(args.min_month)
    except ValueError:
        raise SystemExit("--min-month must be YYYY-MM-DD")
    min_month = _month_first(min_month)

    app = create_app()
    with app.app_context():
        run_bulk_import(
            workbooks_dir=Path(args.workbooks_dir),
            min_month=min_month,
            commit=bool(args.commit),
            allow_completed_runs=bool(args.allow_completed_runs),
            only_workbook=args.only_workbook,
            max_sheets=args.max_sheets,
        )


if __name__ == "__main__":
    main()

